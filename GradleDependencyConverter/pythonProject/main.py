import sys
import os
import re
import argparse
import pandas as pd
import zipfile
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class DependenciesHandler:
    REGEX_PATTERN = re.compile(r'\+--- |\\--- |\|    ')

    def __init__(self, deduplicate):
        self.deduplicate = deduplicate

    @staticmethod
    def split_num_str(s):
        """Split string into numeric and non-numeric part"""
        match = re.match(r"([a-z\-]+)([0-9\.]*)", s, re.I)
        return match.groups() if match else (s, "0")

    @staticmethod
    def sort_func(s):
        """Function to sort strings with numeric endings"""
        str_part, num_part = DependenciesHandler.split_num_str(s)
        return str_part, [int(i) for i in num_part.split('.') if i.isdigit()]

    def flatten_dependencies(self, file_path, output_path, excel):
        lines = self._read_file(file_path)

        result_lines = [
            self._process_line(line) for line in lines if self.REGEX_PATTERN.match(line)
        ]
        result_lines.sort(key=self.sort_func)

        if self.deduplicate:
            result_lines = list(set(result_lines))
            result_lines.sort(key=self.sort_func)

        self._write_output(output_path, result_lines, excel)

    @staticmethod
    def _read_file(file_path):
        try:
            with open(file_path, 'r') as f:
                return f.readlines()
        except Exception as e:
            print(f"An error occurred while reading the file {file_path}: {str(e)}")
            sys.exit(1)

    @staticmethod
    def _process_line(line):
        line = line.split(':', 1)[-1]
        line = re.sub(r'\s*\(\*\)|\s*\([0-9]+\)', '', line)
        return line.strip()

    @staticmethod
    def _write_output(output_path, data, excel):
        try:
            if excel:
                df = pd.DataFrame(data, columns=["dependencies"])
                df.to_excel(output_path, index=False)
            else:
                with open(output_path, 'w') as f:
                    for line in data:
                        f.write(line + '\n')
        except Exception as e:
            print(f"An error occurred while writing the output file {output_path}: {str(e)}")
            sys.exit(1)

    def merge_excel_from_zip(self, zip_path, output_path):
        with zipfile.ZipFile(zip_path, 'r') as myzip:
            excel_files = [file for file in myzip.namelist() if file.endswith('.xlsx')]
            workbook = Workbook()
            workbook.remove(workbook.active)

            total_dependencies = self._process_excel_files(myzip, excel_files, workbook)

            self._create_total_sheet(total_dependencies, workbook)
            workbook.save(output_path)

    def _process_excel_files(self, myzip, excel_files, workbook):
        total_dependencies = []
        for excel_file in excel_files:
            with myzip.open(excel_file) as myfile:
                df = pd.read_excel(myfile)
                total_dependencies.extend(df['dependencies'].tolist())

                if self.deduplicate:
                    df.drop_duplicates(inplace=True)

                df.sort_values(by=['dependencies'], inplace=True)

                ws = workbook.create_sheet(title=os.path.splitext(excel_file)[0])
                for row in dataframe_to_rows(df, index=False, header=True):
                    ws.append(row)

        return total_dependencies

    def _create_total_sheet(self, total_dependencies, workbook):
        if self.deduplicate:
            total_dependencies = list(set(total_dependencies))

        total_dependencies.sort()

        total_df = pd.DataFrame(total_dependencies, columns=["dependencies"])
        total_ws = workbook.create_sheet(title='total')

        for row in dataframe_to_rows(total_df, index=False, header=True):
            total_ws.append(row)

    def handle_total_sheet(self, file_path):
        df = pd.read_excel(file_path, sheet_name='total')

        if self.deduplicate:
            df.drop_duplicates(inplace=True)
            df.sort_values(by=['dependencies'], ascending=False, inplace=True)

        df.to_excel(file_path, sheet_name='total', index=False)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Flatten dependencies tree and merge excel files from a zip file.')
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('-f', '--file', type=str, help='The path to the dependencies file')
    group.add_argument('-mzip', '--merge_zip', type=str, help='The path to the input zip file containing Excel files to merge')
    parser.add_argument('-txt', type=str, help='The path to the output txt file. It outputs the flattened dependencies in a txt file in ascending order. Cannot be used with -mzip/-ozip.')
    parser.add_argument('-exc', type=str, help='The path to the output Excel file. It outputs the flattened dependencies in an Excel file in ascending order under the "dependencies" header. Cannot be used with -mzip/-ozip.')
    parser.add_argument('-ozip', type=str, help='The path to the output Excel file after merging Excel files from the input zip file. It merges all Excel files in the zip into one Excel file, each in a separate sheet named by the original file name. Can only be used with -mzip.')
    parser.add_argument('-dd', '--deduplicate', action='store_true', help='If set, remove duplicate dependencies in the output.')
    parser.add_argument('-dt', '--deduplicate_total', action='store_true', help='If set, remove duplicate dependencies in the total sheet and sort in descending order.')

    args = parser.parse_args()
    handler = DependenciesHandler(args.deduplicate)

    if args.file:
        if args.txt:
            handler.flatten_dependencies(args.file, args.txt, excel=False)
        elif args.exc:
            handler.flatten_dependencies(args.file, args.exc, excel=True)
        else:
            print("Please provide either -txt or -exc with -f")
            sys.exit(1)

    elif args.merge_zip:
        if args.ozip:
            handler.merge_excel_from_zip(args.merge_zip, args.ozip)
            if args.deduplicate_total:
                handler.handle_total_sheet(args.ozip)
        else:
            print("Please provide -ozip with -mzip")
            sys.exit(1)